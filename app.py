import streamlit as st
import pandas as pd
import numpy as np
from ortools.sat.python import cp_model
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
import io
import collections
import re

# ==========================================
# ⚙️ 設定・定数
# ==========================================
st.set_page_config(layout="wide", page_title="中学校時間割作成システム")

# 主要5教科と技能4教科の定義
MAJOR_SUBJECTS = ['国語', '社会', '数学', '理科', '英語']
SKILL_SUBJECTS = ['音楽', '美術', '体育', '技術', '家庭科', '技術家庭']
# 午前中に配置したい教科
PRIORITIZE_AM_SUBJECTS = ['数学', '英語', '国語']

# 表記ゆれ吸収
NAME_CORRECTIONS = {
    "ニシダ": "ニシタ",
    "オオシマ": "オシマ",
}

# ==========================================
# 🛠️ ヘルパー関数
# ==========================================
def clean_name(name):
    """名前の空白除去と表記ゆれ補正"""
    if pd.isna(name) or name == "":
        return ""
    name = str(name).replace(" ", "").replace("　", "")
    return NAME_CORRECTIONS.get(name, name)

def find_col(df, keywords):
    """列名をあいまい検索"""
    for col in df.columns:
        for kw in keywords:
            if kw in col:
                return col
    return None

def format_cell_text(class_name, subject_name):
    """Excelセル内の表記短縮 (例: 1-1数学 -> 11)"""
    if subject_name in ['総合', '道徳', '学活']: return subject_name
    short_class = class_name.replace('-', '')
    if '音美' in subject_name: return f"★{short_class}"
    return short_class

def parse_manual_overrides(text):
    """手動ピン留めテキストを解析"""
    overrides = []
    if not text: return overrides
    for line in text.split('\n'):
        parts = [p.strip() for p in line.split(',')]
        if len(parts) >= 4:
            # 教員orクラス, 曜日, 限, 教科
            overrides.append({'target': parts[0], 'day': parts[1], 'period': int(parts[2]), 'subj': parts[3]})
    return overrides

# ==========================================
# 📊 Excel生成ロジック (マトリックス形式)
# ==========================================
def generate_excel(df_res, classes, teachers, df_const):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    
    # スタイル定義
    thick = Side(style='thick')
    medium = Side(style='medium')
    thin = Side(style='thin')
    hair = Side(style='hair')
    
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    side_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # 薄い緑

    days = ['月', '火', '水', '木', '金']

    # ---------------------------------------------------------
    # シート1: 教員別 (縦:時間, 横:教員)
    # ---------------------------------------------------------
    ws_t = wb.active
    ws_t.title = "教員別"
    
    # ヘッダー
    ws_t.cell(row=6, column=1, value="曜").fill = header_fill
    ws_t.cell(row=6, column=2, value="限").fill = header_fill
    
    for i, t in enumerate(teachers):
        col = 3 + i
        ws_t.cell(row=6, column=col, value=t).fill = header_fill
        ws_t.column_dimensions[get_column_letter(col)].width = 6

    curr = 7
    for d in days:
        periods = [1,2,3,4,5,6] if d != '金' else [1,2,3,4,5]
        max_p = periods[-1]
        for p in periods:
            # 罫線ロジック (Colab版の再現)
            top = thick if p==1 else (medium if p==5 else thin)
            bottom = thick if p==max_p else (medium if p==4 else thin)
            
            # 左サイド (曜日・限)
            c_day = ws_t.cell(row=curr, column=1, value=d if p==1 else "")
            c_day.fill = side_fill
            c_day.border = Border(top=top, bottom=bottom, left=thick, right=thin)
            c_day.alignment = align_center
            
            c_p = ws_t.cell(row=curr, column=2, value=p)
            c_p.fill = side_fill
            c_p.border = Border(top=top, bottom=bottom, left=thin, right=thin)
            c_p.alignment = align_center
            
            # データ埋め込み
            for i, t in enumerate(teachers):
                cell = ws_t.cell(row=curr, column=3+i)
                cell.border = Border(top=top, bottom=bottom, left=hair, right=hair)
                cell.alignment = align_center
                
                # 授業検索
                matches = df_res[(df_res['曜日']==d) & (df_res['限']==p) & (df_res['教員'].str.contains(t, na=False))]
                val = ""
                if not matches.empty:
                    r = matches.iloc[0]
                    val = format_cell_text(r['クラス'], r['教科'])
                else:
                    # 固定リスト(部会など)検索
                    for fix in df_const:
                        if fix['target'] == t and fix['day'] == {'月':0,'火':1,'水':2,'木':3,'金':4}[d] and fix['period'] == p:
                            val = f"【{fix['content']}】"
                            break
                cell.value = val
                if val: cell.font = Font(size=11)
            curr += 1

    # ---------------------------------------------------------
    # シート2: クラス別 (縦:時間, 横:クラス)
    # ---------------------------------------------------------
    ws_c = wb.create_sheet(title="クラス別")
    
    ws_c.cell(row=1, column=1, value="曜").fill = header_fill
    ws_c.cell(row=1, column=2, value="限").fill = header_fill
    
    for i, c in enumerate(classes):
        col = 3 + i
        cell = ws_c.cell(row=1, column=col, value=c)
        cell.fill = header_fill
        cell.alignment = align_center
        ws_c.column_dimensions[get_column_letter(col)].width = 12

    curr = 2
    for d in days:
        periods = [1,2,3,4,5,6] if d != '金' else [1,2,3,4,5]
        max_p = periods[-1]
        for p in periods:
            top = thick if p==1 else (medium if p==5 else thin)
            bottom = thick if p==max_p else (medium if p==4 else thin)
            
            ws_c.cell(row=curr, column=1, value=d if p==1 else "").border = Border(top=top, bottom=bottom, left=thick, right=thin)
            ws_c.cell(row=curr, column=2, value=p).border = Border(top=top, bottom=bottom, left=thin, right=thin)
            
            for i, c in enumerate(classes):
                cell = ws_c.cell(row=curr, column=3+i)
                cell.border = Border(top=top, bottom=bottom, left=thin, right=thin)
                cell.alignment = align_center
                
                matches = df_res[(df_res['曜日']==d) & (df_res['限']==p) & (df_res['クラス']==c)]
                if not matches.empty:
                    r = matches.iloc[0]
                    cell.value = f"{r['教科']}\n{r['教員']}"
                    cell.font = Font(size=9)
            curr += 1
            
    wb.save(output)
    return output.getvalue()


# ==========================================
# 🧩 最適化ロジック (Colab版完全移植)
# ==========================================
def solve_schedule(teachers, req_list, fixed_list, weights, recalc_target_classes, manual_overrides, prev_schedule_df):
    model = cp_model.CpModel()
    DAYS = 5
    days_map = {0:'月', 1:'火', 2:'水', 3:'木', 4:'金'}
    
    # クラス一覧
    classes = sorted(list(set(r['class'] for r in req_list)))
    
    # 変数 X[req_id, day, period]
    X = {}
    class_subjects = collections.defaultdict(list) # クラスごとの授業リスト

    # 1. 変数作成と基本制約
    for r in req_list:
        rid = r['id']
        class_subjects[r['class']].append(r)
        
        slots = []
        for d in range(DAYS):
            p_max = 5 if d == 4 else 6
            for p in range(1, p_max + 1):
                X[(rid, d, p)] = model.NewBoolVar(f'r{rid}_d{d}_p{p}')
                slots.append(X[(rid, d, p)])
        
        # 週コマ数制約
        model.Add(sum(slots) == r['num'])
        
        # 連続制約 (技術家庭など)
        # ★ CSVの設定(continuous) と コマ数(>=2) の両方を満たす場合のみ
        if r['continuous'] and r['num'] >= 2:
            pair_vars = []
            for d in range(DAYS):
                p_max = 5 if d == 4 else 6
                # 昼休み跨ぎ(4-5)禁止のペア
                pairs = [(1,2), (2,3), (3,4)]
                if p_max >= 6: pairs.append((5,6))
                
                for (p1, p2) in pairs:
                    b_pair = model.NewBoolVar(f'pair_{rid}_{d}_{p1}')
                    model.Add(X[(rid, d, p1)] + X[(rid, d, p2)] == 2).OnlyEnforceIf(b_pair)
                    model.Add(X[(rid, d, p1)] + X[(rid, d, p2)] != 2).OnlyEnforceIf(b_pair.Not())
                    pair_vars.append(b_pair)
            
            # 週2コマなら1セット、週4コマなら2セット必要だが、
            # 簡易的に「少なくとも (コマ数//2) セットはある」とする
            model.Add(sum(pair_vars) >= r['num'] // 2)

    # 2. クラス内 重複禁止
    for cls in classes:
        cls_reqs = class_subjects[cls]
        for d in range(DAYS):
            p_max = 5 if d == 4 else 6
            for p in range(1, p_max + 1):
                vars_here = [X[(r['id'], d, p)] for r in cls_reqs if (r['id'], d, p) in X]
                if vars_here:
                    model.Add(sum(vars_here) <= 1)

    # 3. 教員 重複禁止 & 固定リスト
    t_map = {t: [] for t in teachers}
    for r in req_list:
        if r['t1'] in teachers: t_map[r['t1']].append(r)
        if r['t2'] in teachers: t_map[r['t2']].append(r)
    
    for t in teachers:
        # 固定リスト (授業不可)
        for fix in fixed_list:
            if fix['target'] == t:
                d, p = fix['day'], fix['period']
                vars_here = [X[(r['id'], d, p)] for r in t_map[t] if (r['id'], d, p) in X]
                if vars_here:
                    model.Add(sum(vars_here) == 0)
        
        # 重複禁止
        for d in range(DAYS):
            p_max = 5 if d == 4 else 6
            for p in range(1, p_max + 1):
                vars_here = [X[(r['id'], d, p)] for r in t_map[t] if (r['id'], d, p) in X]
                if vars_here:
                    model.Add(sum(vars_here) <= 1)

    # 4. ★重要★ 学年排他 (体育、理科など)
    # Colab版にあったロジックの復活
    grade_reqs = {} # "1": [reqs], "2": [reqs]
    for r in req_list:
        # クラス名 "1-1" -> "1" を抽出
        g = r['class'].split('-')[0]
        if g not in grade_reqs: grade_reqs[g] = []
        grade_reqs[g].append(r)
    
    excl_subjs = ["体育", "理科", "音楽", "美術"]
    for g, reqs in grade_reqs.items():
        for subj_name in excl_subjs:
            # その学年、その教科の授業IDリスト
            target_reqs = [r for r in reqs if subj_name in r['subject'] or "音美" in r['subject']]
            for d in range(DAYS):
                p_max = 5 if d == 4 else 6
                for p in range(1, p_max + 1):
                    vars_here = [X[(r['id'], d, p)] for r in target_reqs if (r['id'], d, p) in X]
                    if vars_here:
                        # 同じ時間に同じ学年で1クラスしか実施できない
                        model.Add(sum(vars_here) <= 1)

    # 5. 音美ルール (Colab版)
    # 「音美」がある日は単独の「音楽」「美術」禁止
    for cls in classes:
        cls_reqs = class_subjects[cls]
        has_onbi = any("音美" in r['subject'] for r in cls_reqs)
        if has_onbi:
            reqs_onbi = [r for r in cls_reqs if "音美" in r['subject']]
            reqs_single = [r for r in cls_reqs if r['subject'] in ["音楽", "美術"]]
            
            for d in range(DAYS):
                p_max = 5 if d == 4 else 6
                # その日に音美があるかフラグ
                is_onbi_day = model.NewBoolVar(f'onbi_day_{cls}_{d}')
                onbi_vars = []
                for p in range(1, p_max + 1):
                    for r in reqs_onbi:
                        if (r['id'], d, p) in X: onbi_vars.append(X[(r['id'], d, p)])
                
                # 音美があればフラグTrue
                model.Add(sum(onbi_vars) >= 1).OnlyEnforceIf(is_onbi_day)
                model.Add(sum(onbi_vars) == 0).OnlyEnforceIf(is_onbi_day.Not())
                
                # フラグTrueなら、単独科目は禁止
                for p in range(1, p_max + 1):
                    for r in reqs_single:
                        if (r['id'], d, p) in X:
                            model.Add(X[(r['id'], d, p)] == 0).OnlyEnforceIf(is_onbi_day)

    # 6. 教員負荷制限 (午前中の会議数に応じて授業制限)
    for t in teachers:
        for d in range(DAYS):
            # 午前中(1-4)の会議数をカウント
            meeting_cnt = 0
            has_shoninzu = False
            for fix in fixed_list:
                if fix['target'] == t and fix['day'] == d and fix['period'] <= 4:
                    meeting_cnt += 1
                    if "少人数" in fix['content']: has_shoninzu = True
            
            # 午前中の授業変数
            am_vars = []
            for p in range(1, 5): # 1-4限
                vars_here = [X[(r['id'], d, p)] for r in t_map[t] if (r['id'], d, p) in X]
                am_vars.extend(vars_here)
            
            if am_vars:
                # 基準: 基本3コマまで。少人数部会なら4コマOK。会議分減らす
                limit_base = 4 if has_shoninzu else 3
                limit_teaching = max(0, limit_base - meeting_cnt)
                model.Add(sum(am_vars) <= limit_teaching)

    # 7. ★再計算ロック機能 (Human-in-the-loop)
    # prev_schedule_df があり、recalc_target_classes が指定されている場合
    if prev_schedule_df is not None and recalc_target_classes:
        # 前回の結果を解析して固定
        # prev_df は縦:時間、横:クラス の形式と想定
        try:
            for index, row in prev_schedule_df.iterrows():
                d_str = row['曜'] if '曜' in row else row['曜日']
                p = int(row['限'])
                d_idx = {'月':0, '火':1, '水':2, '木':3, '金':4}.get(d_str, -1)
                
                if d_idx == -1: continue

                for col_cls in prev_schedule_df.columns:
                    if col_cls not in classes: continue # 列名がクラス名でないならスキップ
                    
                    # 再計算したいクラスなら固定しない
                    if col_cls in recalc_target_classes: continue
                    
                    cell_val = str(row[col_cls])
                    if cell_val == 'nan' or cell_val == '':
                        # 空きコマだった場所 -> 授業を入れない
                        for r in class_subjects[col_cls]:
                            if (r['id'], d_idx, p) in X:
                                model.Add(X[(r['id'], d_idx, p)] == 0)
                    else:
                        # 授業が入っていた場所 -> その授業を固定
                        # セル内容: "国語\n田中" -> 教科名でマッチング
                        val_lines = cell_val.split('\n')
                        subj_name = val_lines[0].strip()
                        
                        # 該当する授業IDを探して固定
                        # (同じ教科が複数ある場合が難しいが、簡易的に最初に見つかったものを固定)
                        for r in class_subjects[col_cls]:
                            if r['subject'] == subj_name:
                                if (r['id'], d_idx, p) in X:
                                    model.Add(X[(r['id'], d_idx, p)] == 1)
                                    # 本当は重複排除が必要だが、簡易実装
                                    break 
        except Exception as e:
            st.warning(f"再計算ロック中にエラーが発生しましたが、続行します: {e}")

    # 8. ★手動ピン留め (Manual Overrides)
    for o in manual_overrides:
        tgt = o['target']
        d_idx = {'月':0, '火':1, '水':2, '木':3, '金':4}.get(o['day'], -1)
        p = o['period']
        subj_name = o['subj']
        
        if d_idx != -1:
            # クラス指定の場合
            if tgt in classes:
                found = False
                for r in class_subjects[tgt]:
                    if r['subject'] == subj_name:
                        if (r['id'], d_idx, p) in X:
                            model.Add(X[(r['id'], d_idx, p)] == 1)
                            found = True
                            break
            # 教員指定の場合 (その先生の該当教科を固定)
            elif tgt in teachers:
                for r in t_map[tgt]:
                    if r['subject'] == subj_name:
                         if (r['id'], d_idx, p) in X:
                            model.Add(X[(r['id'], d_idx, p)] == 1)

    # 9. 目的関数 (Weights適用)
    obj_terms = []
    
    # 基本: 前詰め (WEIGHT_AM_PLACEMENT)
    for (rid, d, p), var in X.items():
        obj_terms.append(var * p * weights['AM_PLACEMENT'])

    # 先生の負担平準化 (TEACHER_LOAD)
    if weights['TEACHER_LOAD'] > 0:
        for t in teachers:
            daily_counts = []
            for d in range(DAYS):
                p_max = 5 if d == 4 else 6
                d_vars = []
                for p in range(1, p_max+1):
                    # その日のその先生の授業変数
                    for r in t_map[t]:
                        if (r['id'], d, p) in X: d_vars.append(X[(r['id'], d, p)])
                
                # その日のコマ数を表す変数
                cnt = model.NewIntVar(0, 6, f'tc_{t}_{d}')
                model.Add(sum(d_vars) == cnt)
                daily_counts.append(cnt)
            
            # 最大 - 最小 をペナルティに
            mx = model.NewIntVar(0, 6, f'tmax_{t}')
            mn = model.NewIntVar(0, 6, f'tmin_{t}')
            model.AddMaxEquality(mx, daily_counts)
            model.AddMinEquality(mn, daily_counts)
            
            # 係数を掛けて追加
            obj_terms.append((mx - mn) * weights['TEACHER_LOAD'])

    model.Minimize(sum(obj_terms))

    # ソルバー実行
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 120.0
    status = solver.Solve(model)

    if status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
        recs = []
        for r in req_list:
            rid = r['id']
            for d in range(DAYS):
                p_max = 5 if d == 4 else 6
                for p in range(1, p_max + 1):
                    if (rid, d, p) in X and solver.Value(X[(rid, d, p)]) == 1:
                        t_str = r['t1']
                        if r['t2']: t_str += f", {r['t2']}"
                        recs.append({
                            '曜日': days_map[d],
                            '限': p,
                            'クラス': r['class'],
                            '教科': r['subject'],
                            '教員': t_str
                        })
        return pd.DataFrame(recs)
    else:
        return None

# ==========================================
# 📱 メインアプリ画面
# ==========================================
st.title("🏫 中学校 時間割作成システム (Colab機能完全版)")

st.sidebar.header("1. データアップロード")
f_teacher = st.sidebar.file_uploader("教員データ", type='csv', key="t")
f_subject = st.sidebar.file_uploader("教科設定", type='csv', key="s")
f_req = st.sidebar.file_uploader("授業データ", type='csv', key="r")
f_fixed = st.sidebar.file_uploader("固定・禁止リスト", type='csv', key="f")
st.sidebar.markdown("---")
f_prev = st.sidebar.file_uploader("🔄 前回データ (再計算用Excel)", type='xlsx', key="prev")

st.sidebar.header("2. こだわり設定 (重み)")
w_load = st.sidebar.slider("先生の負担平準化", 0, 100, 20)
w_am = st.sidebar.slider("主要科目の午前配置", 0, 100, 50)
weights = {'TEACHER_LOAD': w_load, 'AM_PLACEMENT': w_am}

st.sidebar.header("3. 調整・ピン留め")
recalc_str = st.sidebar.text_input("作り直すクラス (例: 1-1, 1-2)", "")
manual_str = st.sidebar.text_area("手動ピン留め (例: 1-1,月,1,国語)", height=100)

if st.sidebar.button("🚀 作成開始"):
    if not all([f_teacher, f_subject, f_req]):
        st.error("⚠️ 必須ファイル（教員、教科、授業）が不足しています。")
    else:
        with st.spinner("計算中..."):
            try:
                # -----------------------
                # データ読み込み処理
                # -----------------------
                # 教員
                df_teacher = pd.read_csv(f_teacher, encoding='utf-8-sig')
                c_name = find_col(df_teacher, ['教員名', '氏名', '名前'])
                df_teacher['教員名'] = df_teacher[c_name].apply(clean_name)
                teachers = df_teacher['教員名'].unique().tolist()
                
                # 教科
                df_subj = pd.read_csv(f_subject, encoding='utf-8-sig')
                c_sname = find_col(df_subj, ['教科名', '教科'])
                c_cont = find_col(df_subj, ['連続'])
                continuous_flags = {}
                for _, row in df_subj.iterrows():
                    s_name = str(row[c_sname]).strip()
                    is_cont = False
                    if c_cont:
                        val = str(row[c_cont])
                        if "〇" in val or "TRUE" in val.upper(): is_cont = True
                    continuous_flags[s_name] = is_cont
                
                # 授業
                df_req = pd.read_csv(f_req, encoding='utf-8-sig')
                c_cls = find_col(df_req, ['クラス'])
                c_sub = find_col(df_req, ['教科'])
                c_t1 = find_col(df_req, ['担当教員', '教員1'])
                c_num = find_col(df_req, ['週コマ', '数'])
                c_t2 = find_col(df_req, ['担当教員2', '教員2'])
                
                req_list = []
                req_id = 0
                for _, row in df_req.iterrows():
                    cls = str(row[c_cls]).strip()
                    subj = str(row[c_sub]).strip()
                    t1 = clean_name(row[c_t1])
                    t2 = clean_name(row[c_t2]) if c_t2 else ""
                    try: num = int(row[c_num])
                    except: continue
                    
                    if num > 0:
                        # ★ 技術家庭科 週1コマなら連続させないロジック
                        is_cont = continuous_flags.get(subj, False)
                        if num < 2: is_cont = False
                        
                        req_list.append({
                            'id': req_id, 'class': cls, 'subject': subj,
                            't1': t1, 't2': t2, 'num': num, 'continuous': is_cont
                        })
                        req_id += 1
                
                # 固定リスト
                fixed_list = []
                if f_fixed:
                    df_fix = pd.read_csv(f_fixed, encoding='utf-8-sig')
                    c_tar = find_col(df_fix, ['対象', '教員'])
                    c_day = find_col(df_fix, ['曜日'])
                    c_per = find_col(df_fix, ['限'])
                    c_con = find_col(df_fix, ['内容'])
                    if c_tar:
                        for _, row in df_fix.iterrows():
                            target = clean_name(row[c_tar])
                            day_str = row[c_day]
                            try: p = int(row[c_per])
                            except: p = 0
                            content = row[c_con] if c_con else "用務"
                            w_map = {'月':0, '火':1, '水':2, '木':3, '金':4}
                            if day_str in w_map and p > 0:
                                fixed_list.append({'target': target, 'day': w_map[day_str], 'period': p, 'content': content})

                # 再計算用データ
                prev_df = None
                recalc_classes = [x.strip() for x in recalc_str.split(',')] if recalc_str else []
                if f_prev:
                    prev_df = pd.read_excel(f_prev, sheet_name='クラス別')

                # 手動ピン留めパース
                manual_overrides = parse_manual_overrides(manual_str)

                # -----------------------
                # 最適化実行
                # -----------------------
                df_result = solve_schedule(teachers, req_list, fixed_list, weights, recalc_classes, manual_overrides, prev_df)
                
                if df_result is not None:
                    st.success("🎉 時間割が完成しました！")
                    excel_data = generate_excel(df_result, sorted(list(set(r['class'] for r in req_list))), teachers, fixed_list)
                    st.download_button("📥 完成Excelをダウンロード", excel_data, "時間割.xlsx")
                else:
                    st.error("❌ 解が見つかりませんでした。条件を緩和するか、ピン留めを見直してください。")

            except Exception as e:
                st.error(f"エラー: {e}")
